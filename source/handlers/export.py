# -*- coding:utf-8 -*-
from collections import defaultdict
from functools import reduce

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import points_to_pixels
from sqlalchemy import func
from tornado.web import Finish

from dal import models
from handlers.base.pub_func import TimeFunc, check_float
from handlers.base.pub_web import StationBaseHandler
from handlers.base.webbase import BaseHandler


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


# 意向单导出
class WishOrderExport(StationBaseHandler):
    def get(self, order_id):
        order = models.WishOrder.get_by_id(self.session, order_id, self.current_station.id)
        if not order:
            self.write("没有找到该意向单")
            raise Finish()

        goods_list = models.WishOrderGoods.get_by_order_id(self.session, order_id)

        # 默认采购员信息
        goods_ids = {goods.goods_id for goods in goods_list}
        default_purchasers = self.session.query(models.StaffGoods, models.Staff, models.AccountInfo) \
            .join(models.Staff, models.Staff.id == models.StaffGoods.staff_id) \
            .join(models.AccountInfo, models.AccountInfo.id == models.Staff.account_id) \
            .filter(models.StaffGoods.goods_id.in_(goods_ids)) \
            .all()
        default_purchaser_dict = {item[0].goods_id: item for item in default_purchasers}

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "意向单"

        # 表头
        worksheet.cell(column=1, row=1, value="嘿嘛{}月{}日各店剩货及意向单".format(order.wish_date.month, order.wish_date.day))
        worksheet.merge_cells("A1:G1")
        worksheet.cell(column=1, row=2, value="序号")
        worksheet.cell(column=2, row=2, value="条码")
        worksheet.cell(column=3, row=2, value="商品名称")
        worksheet.cell(column=4, row=2, value="负责人")
        worksheet.cell(column=5, row=2, value="剩货数量")
        worksheet.cell(column=6, row=2, value="意向数量")
        worksheet.cell(column=7, row=2, value="说明")

        # 意向单商品列表
        for idx, wish_goods in enumerate(goods_list):
            row = 3 + idx
            goods = wish_goods.goods
            purchaser = default_purchaser_dict.get(goods.id)
            purchaser_name = purchaser[1].remarks or purchaser[2].username if purchaser else ""

            worksheet.cell(column=1, row=row, value=idx + 1)
            worksheet.cell(column=2, row=row, value=goods.code or "")
            worksheet.cell(column=3, row=row, value=wish_goods.goods_name or goods.name)
            worksheet.cell(column=4, row=row, value=purchaser_name)
            worksheet.cell(column=5, row=row, value="")
            worksheet.cell(column=6, row=row, value="")
            worksheet.cell(column=7, row=row, value=wish_goods.remarks)

        file_name = "{}各店意向单".format(TimeFunc.time_to_str(order.wish_date, "date"))
        return self.export_xlsx(workbook, file_name)


# 店铺配货单导出
class ShopPackingOrderExport(StationBaseHandler):
    def get(self, order_id):
        demand_order = models.DemandOrder.get_by_id(self.session, order_id)

        if not demand_order:
            self.write("没有找到此订货单")
            raise Finish()

        wish_order = models.WishOrder.get_by_id(self.session, demand_order.wish_order_id)

        if not wish_order:
            self.write("没有找到对应的意向单")
            raise Finish()

        order_goods_list = self.session.query(models.DemandOrderGoods) \
            .filter(models.DemandOrderGoods.demand_order_id == order_id,
                    models.DemandOrderGoods.status == 0) \
            .all()

        # 所有已确认的分车记录
        allocated_amount = self.session.query(models.AllocationOrder.goods_id,
                                              func.sum(models.AllocationOrderGoods.actual_allocated_amount)) \
            .join(models.AllocationOrderGoods, models.AllocationOrder.id == models.AllocationOrderGoods.order_id) \
            .filter(models.AllocationOrder.wish_order_id == wish_order.id,
                    models.AllocationOrder.status == 1,
                    models.AllocationOrderGoods.shop_id == demand_order.shop_id) \
            .group_by(models.AllocationOrder.goods_id) \
            .all()
        allocated_amount_dict = {data[0]: data[1] for data in allocated_amount}

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "店铺配货单"

        # 表头
        title = "{}月{}日 {} 店配货单".format(wish_order.wish_date.month, wish_order.wish_date.day, demand_order.shop.abbreviation)
        worksheet.cell(column=1, row=1, value=title)
        worksheet.merge_cells("A1:F1")
        worksheet.cell(column=1, row=2, value="序号")
        worksheet.cell(column=2, row=2, value="商品名称")
        worksheet.cell(column=3, row=2, value="分店库存")
        worksheet.cell(column=4, row=2, value="订货量")
        worksheet.cell(column=5, row=2, value="实配量")
        worksheet.cell(column=6, row=2, value="备注")

        # 意向单商品列表
        for idx, order_goods in enumerate(order_goods_list):
            row = 3 + idx
            current_storage = check_float(order_goods.current_storage / 100)
            demand_amount = check_float(order_goods.demand_amount / 100)
            demand_remarks = order_goods.remarks

            wish_order_goods = order_goods.wish_order_goods

            worksheet.cell(column=1, row=row, value=idx + 1)
            worksheet.cell(column=2, row=row, value=wish_order_goods.goods_name if wish_order_goods else order_goods.goods.name)
            worksheet.cell(column=3, row=row, value=current_storage)
            worksheet.cell(column=4, row=row, value=demand_amount)
            worksheet.cell(column=5, row=row, value=check_float(allocated_amount_dict.get(order_goods.goods_id, 0) / 100))
            worksheet.cell(column=6, row=row, value=demand_remarks)

        file_name = title
        return self.export_xlsx(workbook, file_name)


# 进货单导出
class SummaryExport(StationBaseHandler):
    @BaseHandler.check_arguments("wish_order_id:int")
    def get(self):
        wish_order_id = self.args["wish_order_id"]

        wish_order = models.WishOrder.get_by_id(self.session, wish_order_id)

        if not wish_order:
            self.write("没有找到对应的意向单")
            raise Finish()

        # 所有订货单
        demand_orders_shops = self.session.query(models.DemandOrder, models.Shop) \
            .join(models.Shop) \
            .filter(models.DemandOrder.wish_order_id == wish_order_id,
                    models.DemandOrder.status == 2) \
            .all()

        # 订货的门店列表，表头门店顺序基于此
        shops = list({shop for o, shop in demand_orders_shops})
        shop_order = {"仓库": -999, "总部": 1, "刘园": 2, "侯台": 3,
                      "咸水沽": 4, "华明": 5,
                      "大港": 6, "杨村": 7, "新立": 8, "大寺": 9, "汉沽": 10,
                      "沧州": 11, "静海": 13, "芦台": 14, "工农村": 15, "唐山": 16, "廊坊": 17,
                      "哈尔滨": 18, "西青道": 19, "双鸭山": 20, "承德": 21,
                      "张胖子": 22, "固安": 23, "燕郊": 24, "胜芳": 25, "蓟县": 26, }
        shops = sorted(shops, key=lambda d: shop_order.get(d.abbreviation, 999))

        # 意向单商品列表，左侧商品列表顺序基于此
        wish_goods_list = self.session.query(models.WishOrderGoods, models.Goods) \
            .join(models.Goods) \
            .filter(models.WishOrderGoods.wish_order_id == wish_order_id,
                    models.WishOrderGoods.status >= 0) \
            .order_by(models.WishOrderGoods.status.asc(),
                      models.WishOrderGoods.priority.asc()) \
            .all()

        # 所有订货单商品
        demand_order_ids = {o.id for o, s in demand_orders_shops}
        demand_goods_list = self.session.query(models.DemandOrderGoods) \
            .filter(models.DemandOrderGoods.demand_order_id.in_(demand_order_ids)) \
            .all()
        demand_order_dict = {o.id: o for o, s in demand_orders_shops}
        demand_goods_dict = {
            "{}:{}".format(g.wish_order_goods_id,
                           demand_order_dict.get(g.demand_order_id).shop_id if g.demand_order_id else -1)
            : g
            for g in demand_goods_list
        }

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "母单"
        worksheet.freeze_panes = "F4"

        alignment_center = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        fill_current_storage = PatternFill("solid", fgColor="A6A6A6")
        fill_ambiguous_demand_amount = PatternFill("solid", fgColor="92D050")
        font_title = Font(name="华文楷体", size=18, bold=True)
        font_head = Font(name="宋体", size=11)
        font_head_accent = Font(name="宋体", size=11, color="FF0000")
        font_shop_name = Font(name="微软雅黑", size=11, bold=True)
        font_goods_name = Font(name="宋体", size=11)
        font_goods_name_accent = Font(name="宋体", size=11, color="FF0000")
        font_sum_demand_amount = Font(name="微软雅黑", size=11, bold=True)
        font_demand_amount = Font(name="微软雅黑", size=20, bold=True, color="FF0000")
        font_current_storage = Font(name="微软雅黑", size=10)

        # 库存列宽
        worksheet.column_dimensions[get_column_letter(1)].width = points_to_pixels(3.83)
        # 商品编码列宽
        worksheet.column_dimensions[get_column_letter(2)].width = points_to_pixels(0)
        # 商品名列宽
        worksheet.column_dimensions[get_column_letter(3)].width = points_to_pixels(16.33)
        # 负责人列宽
        worksheet.column_dimensions[get_column_letter(4)].width = points_to_pixels(1.83)
        # 总数量列宽
        worksheet.column_dimensions[get_column_letter(5)].width = points_to_pixels(3.83)
        # 门店名行高
        worksheet.row_dimensions[2].height = points_to_pixels(28)
        # 意向数量表头行高
        worksheet.row_dimensions[3].height = points_to_pixels(13)

        # 标题
        title = "嘿嘛{}月{}日各店剩货及意向单".format(wish_order.wish_date.month, wish_order.wish_date.day)
        cell_title = worksheet.cell(column=1, row=1, value=title)
        cell_title.font = font_title
        cell_title.alignment = alignment_center
        worksheet.merge_cells(start_row=1, end_row=1,
                              start_column=1, end_column=5 + 2 * len(shops))  # 合并标题
        style_range(worksheet, "{}{}:{}{}".format(get_column_letter(1), 1, get_column_letter(5 + 2 * len(shops)), 1), border=border)
        # 表头
        cell_head_storage = worksheet.cell(column=1, row=2, value="库存")
        cell_head_storage.font = font_head
        cell_head_storage.alignment = alignment_center
        worksheet.merge_cells("A2:A3")
        style_range(worksheet, "A2:A3", border=border)
        cell_head_code = worksheet.cell(column=2, row=2, value="商品编码")
        cell_head_code.font = font_head
        cell_head_code.alignment = alignment_center
        worksheet.merge_cells("B2:B3")
        style_range(worksheet, "B2:B3", border=border)
        cell_head_goods_name = worksheet.cell(column=3, row=2, value="商品名称")
        cell_head_goods_name.font = font_head
        cell_head_goods_name.alignment = alignment_center
        cell_head_goods_name.border = border
        worksheet.merge_cells("C2:C3")
        style_range(worksheet, "C2:C3", border=border)
        cell_head_purchaser = worksheet.cell(column=4, row=2, value="负责人")
        cell_head_purchaser.font = font_head
        cell_head_purchaser.alignment = alignment_center
        cell_head_purchaser.border = border
        worksheet.merge_cells("D2:D3")
        style_range(worksheet, "D2:D3", border=border)
        cell_head_sum = worksheet.cell(column=5, row=2, value="总数量")
        cell_head_sum.font = font_head
        cell_head_sum.alignment = alignment_center
        cell_head_sum.border = border
        cell_head_sum_demand_amount = worksheet.cell(column=5, row=3, value="意向数量")
        cell_head_sum_demand_amount.font = font_head
        cell_head_sum_demand_amount.alignment = alignment_center
        cell_head_sum_demand_amount.border = border

        # 填门店名
        for i, shop in enumerate(shops):
            # 跳过货品信息和总订货量列
            base_column = 6 + i * 2
            # 跳过标题
            base_row = 2
            # 门店名表头
            cell_shop_name = worksheet.cell(column=base_column, row=base_row, value=shop.abbreviation)
            cell_shop_name.font = font_shop_name
            cell_shop_name.alignment = alignment_center
            cell_shop_name.border = border
            worksheet.merge_cells(start_row=base_row, end_row=base_row, start_column=base_column, end_column=base_column + 1)
            style_range(worksheet, "{}{}:{}{}".format(get_column_letter(base_column), base_row,
                                                      get_column_letter(base_column + 1), base_row), border=border)
            # 剩货数量表头
            cell_head_current_storage = worksheet.cell(column=base_column, row=base_row + 1, value="剩货数量")
            cell_head_current_storage.font = font_head
            cell_head_current_storage.alignment = alignment_center
            cell_head_current_storage.border = border
            # 意向数量表头
            cell_head_demand_amount = worksheet.cell(column=base_column + 1, row=base_row + 1, value="意向数量")
            cell_head_demand_amount.font = font_head_accent
            cell_head_demand_amount.alignment = alignment_center
            cell_head_demand_amount.border = border

            # 列宽
            worksheet.column_dimensions[get_column_letter(base_column)].width = points_to_pixels(0.91)
            worksheet.column_dimensions[get_column_letter(base_column + 1)].width = points_to_pixels(3.83)

        # 填货品信息
        ambiguous = False
        ambiguous_row = -1
        ambiguous_sum_dict = defaultdict(float)
        for i, (wish_goods, goods) in enumerate(wish_goods_list):
            # 跳过表头
            base_row = 4 + i

            # 不保证有货行
            if wish_goods.status > 0 and not ambiguous:
                ambiguous = True
                ambiguous_row = base_row
                cell_head_ambiguous_goods_name = worksheet.cell(column=3, row=base_row, value="下方产品不保证有货")
                cell_head_ambiguous_goods_name.font = font_goods_name_accent
                cell_head_ambiguous_goods_name.alignment = alignment_center
                cell_head_ambiguous_goods_name.border = border

            # 被不保证有货行下推一行
            if ambiguous:
                base_row += 1

            # 库存
            cell_storage = worksheet.cell(column=1, row=base_row, value=check_float(goods.stock / 100))
            cell_storage.alignment = alignment_center
            cell_storage.border = border
            cell_storage.font = font_goods_name

            # 商品编码
            cell_code = worksheet.cell(column=2, row=base_row, value=goods.code)
            cell_code.alignment = alignment_center
            cell_code.border = border
            cell_code.font = font_goods_name

            # 商品名
            cell_goods_name = worksheet.cell(column=3, row=base_row, value=wish_goods.goods_name or goods.name)
            cell_goods_name.alignment = alignment_center
            cell_goods_name.border = border
            if wish_goods.goods_name_modified:
                cell_goods_name.font = font_goods_name_accent
            else:
                cell_goods_name.font = font_goods_name

            # 负责人
            cell_purchaser = worksheet.cell(column=4, row=base_row, value="")
            cell_purchaser.alignment = alignment_center
            cell_purchaser.border = border
            cell_purchaser.font = font_goods_name

            demand_sum = 0
            for j, shop in enumerate(shops):
                demand_goods = demand_goods_dict.get("{}:{}".format(wish_goods.id, shop.id))
                # 跳过货品信息和总订货量列
                base_column = 6 + j * 2
                # 门店库存
                current_storage = check_float(demand_goods.current_storage / 100) if demand_goods else 0
                cell_current_storage = worksheet.cell(column=base_column, row=base_row, value=current_storage)
                cell_current_storage.font = font_current_storage
                cell_current_storage.alignment = alignment_center
                cell_current_storage.border = border
                cell_current_storage.fill = fill_current_storage
                # 门店订货量
                if demand_goods:
                    if demand_goods.modified_demand_amount is None:
                        demand_amount = demand_goods.demand_amount
                    else:
                        demand_amount = demand_goods.modified_demand_amount
                    demand_amount = check_float(demand_amount / 100)
                else:
                    demand_amount = 0
                cell_demand_amount = worksheet.cell(column=base_column + 1, row=base_row, value=demand_amount)
                cell_demand_amount.font = font_demand_amount
                cell_demand_amount.alignment = alignment_center
                cell_demand_amount.border = border

                # 该商品各店总订货量
                demand_sum += demand_amount
                # 各店不保证有货商品总订货量
                if wish_goods.status > 0:
                    ambiguous_sum_dict[shop.id] += demand_amount

            # 总意向数量
            cell_sum_demand_amount = worksheet.cell(column=5, row=base_row, value=check_float(demand_sum))
            cell_sum_demand_amount.font = font_sum_demand_amount
            cell_sum_demand_amount.alignment = alignment_center
            cell_sum_demand_amount.border = border

        # 不保证有货商品行的信息
        if ambiguous_row >= 0:
            for j, shop in enumerate(shops):
                # 跳过货品信息和总订货量列，取意向数量列
                base_column = 6 + j * 2 + 1
                demand_sum = check_float(ambiguous_sum_dict.get(shop.id, 0))
                cell_sum_ambiguous_demand_amount = worksheet.cell(column=base_column, row=ambiguous_row, value=demand_sum)
                cell_sum_ambiguous_demand_amount.font = font_demand_amount
                cell_sum_ambiguous_demand_amount.alignment = alignment_center
                cell_sum_ambiguous_demand_amount.border = border
                cell_sum_ambiguous_demand_amount.fill = fill_ambiguous_demand_amount
            # 所有店不保证有货商品订货量总和
            total_ambiguous_sum = reduce(lambda s, i: s + i, ambiguous_sum_dict.values(), 0)
            cell_sum_ambiguous_demand_amount = worksheet.cell(column=5, row=ambiguous_row, value=check_float(total_ambiguous_sum))
            cell_sum_ambiguous_demand_amount.font = font_sum_demand_amount
            cell_sum_ambiguous_demand_amount.alignment = alignment_center
            cell_sum_ambiguous_demand_amount.border = border

        file_name = title
        return self.export_xlsx(workbook, file_name)
