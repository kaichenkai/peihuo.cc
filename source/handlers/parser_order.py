#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
from io import BytesIO
from openpyxl import load_workbook
from dal import models
from handlers.base.pub_web import StationBaseHandler
from handlers.base.webbase import BaseHandler
from dal.constants import PURCHASER_MAP


# 订单解析
class ParserOrder(StationBaseHandler):
    @BaseHandler.check_arguments("action:str")
    def post(self):
        action = self.args["action"]
        files = self.request.files
        if not files:
            return self.send_fail("请选择上传文件")
        if not isinstance(files, dict):
            return self.send_fail("文件上传参数有误")
        file_list = next(iter(files.values()))
        file = file_list[0]

        file_name = file["filename"]
        content = file["body"]

        # 检验文件的后缀名是否正确
        suffix_name = file_name.split(".")[-1]
        if suffix_name not in ('xlsx', 'xlsm', 'xltx', 'xltm'):
            self.delete_file(file_name)
            return self.send_fail("文件格式有误")
        # 从内存中获取字节文件
        wb = load_workbook(filename=BytesIO(content))
        sheet = wb.active
        if not sheet:
            return self.send_fail("获取文件数据失败")

        if action == "wish_order_parser":
            valid, message, wish_order_id = self.wish_order_parser(file_name, sheet)
            self.delete_file(file_name)
            if not valid:
                return self.send_fail(message)
            else:
                return self.send_success(wish_order_id=wish_order_id)
        elif action == "demand_order_parser":
            valid, message = self.demand_order_parser(file_name, sheet)
            self.delete_file(file_name)
            if not valid:
                return self.send_fail(message)
            else:
                return self.send_success()
        else:
            return self.send_fail("不支持的操作类型")

    @BaseHandler.check_arguments("wish_date:date")
    def wish_order_parser(self, file_name, sheet):
        wish_date = self.args["wish_date"]
        station_id = self.current_station.id
        creator_id = self.current_user.id

        # file_name = '2018.11.17各店意向单(1)(1).xlsx'
        # wb = load_workbook(file_name)
        # sheet = wb.active

        # 先搞个意向单
        wish_order = self.session.query(models.WishOrder) \
            .filter(models.WishOrder.wish_date == wish_date,
                    models.WishOrder.station_id == station_id) \
            .first()
        if not wish_order:
            wish_order = models.WishOrder(
                wish_date=wish_date,
                status=2,  # 直接是提交状态，要改再说
                station_id=station_id,
                creator_id=creator_id,
            )
            self.session.add(wish_order)
            self.session.flush()

        # 每次直接把之前的意向单商品全删了重建
        self.session.query(models.WishOrderGoods) \
            .filter(models.WishOrderGoods.wish_order_id == wish_order.id) \
            .update({"status": -1})
        self.session.flush()

        goods_status = 0  # 0-有货 1-缺货 2-没货
        for i in range(4, 999999):
            # 取出表里一行的数据
            no = sheet["A{}".format(i)].value
            code = sheet["B{}".format(i)].value
            name = sheet["C{}".format(i)].value
            purchaser = sheet["D{}".format(i)].value
            remarks = sheet["G{}".format(i)].value
            # 如果序号和商品名称都没有，则表示数据导入完成
            if not no and not name:
                break
            if name == '下方产品不保证有货':
                goods_status = 1
                continue
            if remarks and '没有了' in remarks:
                goods_status = 2

            # print('{} {} {} {} {}'.format(no, code, name, purchaser, remarks))

            # 用名字把商品查出来，没有就建一个
            goods = self.session.query(models.Goods) \
                .filter(models.Goods.name == name,
                        models.Goods.station_id == station_id,
                        models.Goods.status == 0) \
                .first()
            if not goods:
                # print('-------------- {} 还没有，建一个 --------------'.format(name))
                goods = models.Goods(
                    name=name,
                    station_id=station_id,
                    creator_id=creator_id,
                )
                self.session.add(goods)
                self.session.flush()
            goods.code = code

            # 拿本行数据建一个意向单商品
            wish_goods = models.WishOrderGoods(
                remarks=remarks,
                status=goods_status,
                goods_id=goods.id,
                wish_order_id=wish_order.id,
                goods_name=goods.name,
                priority=i,
            )
            self.session.add(wish_goods)
            self.session.flush()

            if purchaser:
                purchaser_id = PURCHASER_MAP.get(purchaser)

                if not purchaser_id:
                    return False, "{} 没有对应的采购员".format(purchaser), 0

                default_purchaser = self.session.query(models.StaffGoods) \
                    .filter(models.StaffGoods.goods_id == goods.id) \
                    .first()
                if not default_purchaser:
                    default_purchaser = models.StaffGoods(
                        goods_id=goods.id,
                        staff_id=purchaser_id,
                    )
                    self.session.add(default_purchaser)
                    self.session.flush()
                else:
                    default_purchaser.staff_id = purchaser_id
                    self.session.flush()

        self.session.commit()
        return True, "", wish_order.id

    @BaseHandler.check_arguments("wish_order_id:int")
    def demand_order_parser(self, file_name, sheet):
        wish_order_id = self.args["wish_order_id"]

        # file_name = '2018.12.4西青道店意向单.xlsx'
        # wb = load_workbook(file_name)
        # sheet = wb.active

        # 从文件名里匹配门店名
        matcher = re.search('\d([^\d]+)店', file_name)
        if matcher:
            shop_name = matcher.group(1)
        else:
            # raise Exception('{} 文件里没找到有效店名')
            return False, "{} 文件里没找到有效店名".format(file_name)

        # 把意向单拿出来
        wish_order = models.WishOrder.get_by_id(self.session, wish_order_id)
        if not wish_order:
            # raise Exception('先建个意向单')
            return False, "需要先建个意向单"
        station_id = wish_order.station_id
        creator_id = wish_order.creator_id

        # 从文件名里找一个 Shop，没有就建一个
        shop = self.session.query(models.Shop) \
            .filter(models.Shop.abbreviation == shop_name,
                    models.Shop.station_id == station_id,
                    models.Shop.status == 0) \
            .first()
        if not shop:
            # print('-------------- {} 店还没有，建一个 --------------'.format(shop_name))
            shop = models.Shop(
                abbreviation=shop_name,
                station_id=station_id,
                creator_id=creator_id,
            )
            self.session.add(shop)
            self.session.flush()

        # 建一个订货单
        demand_order = self.session.query(models.DemandOrder) \
            .filter(models.DemandOrder.wish_order_id == wish_order_id,
                    models.DemandOrder.shop_id == shop.id) \
            .first()
        if not demand_order:
            # print('-------------- {} 店还没订货，建一个订货单 --------------'.format(shop_name))
            demand_order = models.DemandOrder(
                wish_order_id=wish_order_id,
                shop_id=shop.id,
                creator_id=creator_id,
                status=2,  # 直接加进汇总单里
            )
            self.session.add(demand_order)
            self.session.flush()

        # 每次直接把之前的订货单商品全删了重建
        self.session.query(models.DemandOrderGoods) \
            .filter(models.DemandOrderGoods.demand_order_id == demand_order.id,
                    models.DemandOrderGoods.status == 0) \
            .update({"status": -1})
        self.session.flush()

        # 把意向单商品都取出来
        wish_goods_list = self.session.query(models.WishOrderGoods) \
            .join(models.WishOrder, models.WishOrder.id == models.WishOrderGoods.wish_order_id) \
            .filter(models.WishOrder.id == wish_order_id,
                    models.WishOrderGoods.status >= 0) \
            .all()
        wish_goods_dict = {goods.goods_id: goods for goods in wish_goods_list}

        for i in range(4, 999999):
            # 取出表里一行的数据
            no = sheet["A{}".format(i)].value
            code = sheet["B{}".format(i)].value
            name = sheet["C{}".format(i)].value
            purchaser = sheet["D{}".format(i)].value
            storage = sheet["E{}".format(i)].value
            demand_amount = sheet["F{}".format(i)].value

            # 如果序号和商品名称都没有，则表示数据导入完成
            if not no and not name:
                break
            if name == '下方产品不保证有货':
                continue

            if storage and not isinstance(storage, float) and not isinstance(storage, int):
                return False, "{} 的库存量 {} 不是纯数字，手动改改".format(name, storage)

            if demand_amount and not isinstance(demand_amount, float) and not isinstance(demand_amount, int):
                return False, "{} 的订货量 {} 不是纯数字，手动改改".format(name, demand_amount)

            # 用名字把商品查出来
            goods = self.session.query(models.Goods) \
                .filter(models.Goods.name == name,
                        models.Goods.station_id == station_id) \
                .first()
            if not goods:
                return False, "没找到 {}，先手动建好加到意向单里吧".format(name)

            wish_goods = wish_goods_dict.get(goods.id)
            if not wish_goods:
                return False, "意向单里没有 {}，先去手动加到意向单里".format(name)

            # 拿本行数据建一个订货单商品
            demand_goods = models.DemandOrderGoods(
                current_storage=round(storage * 100) if storage else 0,
                demand_amount=round(demand_amount * 100) if demand_amount else 0,
                goods_id=goods.id,
                demand_order_id=demand_order.id,
                wish_order_goods_id=wish_goods.id,
            )
            self.session.add(demand_goods)
            self.session.flush()

        self.session.commit()
        return True, ""

    def delete_file(self, file_name):
        # 删除临时保存的文件
        delete_file = "./scripts/{0}".format(file_name)
        if os.path.exists(delete_file):
            os.remove(delete_file)
